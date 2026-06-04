import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from api.models import Student

class Command(BaseCommand):
    help = 'Imports students from an Excel file into the Neon database' #[cite: 19]

    def add_arguments(self, parser):
        # This allows us to pass the filename in the terminal[cite: 19]
        parser.add_argument('excel_file', type=str, help='Path to the excel file') #[cite: 19]

    def handle(self, *args, **kwargs):
        file_path = kwargs['excel_file'] #[cite: 19]
        
        try:
            workbook = openpyxl.load_workbook(file_path) #[cite: 19]
            sheet = workbook.active #[cite: 19]
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error opening file: {e}")) #[cite: 19]
            return

        success_count = 0 #[cite: 19]
        
        # iter_rows(min_row=2) skips the header row[cite: 19]
        for row in sheet.iter_rows(min_row=2, values_only=True): #[cite: 19]
            # Assuming Column A is Full Name, Column B is Student ID[cite: 19]
            full_name = str(row[0]).strip() if row[0] else '' #[cite: 19]
            student_id = str(row[1]).strip() if row[1] else '' #[cite: 19]

            if not full_name or not student_id:
                continue

            # Split the full name into First Name and Last Name[cite: 19]
            name_parts = full_name.split(' ', 1) #[cite: 19]
            first_name = name_parts[0] #[cite: 19]
            last_name = name_parts[1] if len(name_parts) > 1 else '' #[cite: 19]

            # Check if student already exists to prevent crashing[cite: 19]
            if not User.objects.filter(username=student_id).exists(): #[cite: 19]
                
                # Construct the university email format dynamically
                student_email = f"{student_id.lower()}@xmu.edu.my"

                # 1. Create the User with the default password and the new email
                user = User.objects.create_user(
                    username=student_id,
                    first_name=first_name,
                    last_name=last_name,
                    email=student_email,
                    password='samspass123' #[cite: 19]
                )
                
                # 2. Create the linked Student profile[cite: 19]
                Student.objects.create(
                    user=user,  #[cite: 19]
                    has_changed_password=False #[cite: 19]
                )
                
                self.stdout.write(self.style.SUCCESS(f'✅ Created: {full_name} ({student_id})')) #[cite: 19]
                success_count += 1 #[cite: 19]
            else:
                self.stdout.write(self.style.WARNING(f'⚠️ Skipped: {student_id} already exists.')) #[cite: 19]

        self.stdout.write(self.style.SUCCESS(f'\n🎉 Import Complete! Added {success_count} students to the database.')) #[cite: 19]